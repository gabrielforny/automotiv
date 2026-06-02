from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from src.config import ExcelConfig


@dataclass(frozen=True)
class BudgetItem:
    row_number: int
    code_or_reference: str
    quantity: int | float | str
    item: str | None = None


def _normalize_header(value: object) -> str:
    """Normaliza cabeçalhos vindos do Excel.

    A planilha do cliente pode vir com quebras de linha, barras, acentos e
    múltiplos blocos repetidos de colunas. Por isso a comparação precisa ser
    tolerante e não pode depender do texto exatamente igual ao config.yaml.
    """
    text = str(value or "").strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _is_empty(value: object) -> bool:
    return value is None or str(value).strip() == ""


def _matches_header(header: str, expected: str, aliases: list[str] | None = None) -> bool:
    expected_norm = _normalize_header(expected)
    candidates = [expected_norm, *[_normalize_header(alias) for alias in aliases or []]]
    candidates = [candidate for candidate in candidates if candidate]

    return any(
        header == candidate or candidate in header or header in candidate
        for candidate in candidates
    )


def _to_clean_text(value: object) -> str:
    if value is None:
        return ""
    # Excel pode retornar números inteiros como float. Evita gerar código "47688952.0".
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _discover_column_groups(sheet, config: ExcelConfig) -> list[tuple[int | None, int, int]]:
    """Descobre todos os blocos ITEM/CÓDIGO/QUANTIDADE da linha de cabeçalho.

    A planilha enviada não possui apenas um bloco de colunas. Ela repete a
    estrutura 3 vezes na mesma linha:

    A:ITEM | B:CÓDIGO... | E:QUANTIDADE
    H:ITEM | I:CÓDIGO... | L:QUANTIDADE
    O:ITEM | P:CÓDIGO... | S:QUANTIDADE

    O bug anterior acontecia porque o dicionário de headers sobrescrevia as
    colunas repetidas e ficava somente com o último bloco, que estava vazio.
    """
    code_aliases = [
        "CÓDIGO REFERÊNCIA PARTE NUMBER",
        "CODIGO REFERENCIA PARTE NUMBER",
        "CODIGO REFERENCIA PARTE NUMBER PN",
        "CÓDIGO REFERÊNCIA PARTE NUMBER PN",
        "CODIGO / REFERENCIA / PARTE NUMBER",
        "CÓDIGO / REFERÊNCIA / PARTE NUMBER",
        "REFERENCIA",
        "REFERÊNCIA",
        "PARTE NUMBER",
    ]
    qty_aliases = ["QTD", "QTDE", "QUANT", "QUANTIDADE"]
    item_aliases = ["ITEM"]

    header_cells = []
    for cell in sheet[config.header_row]:
        if _is_empty(cell.value):
            continue
        header_cells.append((cell.column, _normalize_header(cell.value), cell.value))

    code_columns = [
        column
        for column, normalized, _original in header_cells
        if _matches_header(normalized, config.columns.code_or_reference, code_aliases)
    ]

    groups: list[tuple[int | None, int, int]] = []
    for code_col in code_columns:
        next_code_cols = [col for col in code_columns if col > code_col]
        group_end = min(next_code_cols) if next_code_cols else sheet.max_column + 1

        qty_col = None
        for column, normalized, _original in header_cells:
            if code_col < column < group_end and _matches_header(normalized, config.columns.quantity, qty_aliases):
                qty_col = column
                break

        item_col = None
        for column, normalized, _original in reversed(header_cells):
            if column < code_col and _matches_header(normalized, config.columns.item or "ITEM", item_aliases):
                # Pega o ITEM mais próximo à esquerda dentro do mesmo bloco.
                if code_col - column <= 4:
                    item_col = column
                    break

        if qty_col:
            groups.append((item_col, code_col, qty_col))

    if groups:
        return groups

    # Fallback para planilhas simples com um único bloco.
    headers = {
        _normalize_header(cell.value): cell.column
        for cell in sheet[config.header_row]
        if not _is_empty(cell.value)
    }

    code_col = next(
        (
            column
            for header, column in headers.items()
            if _matches_header(header, config.columns.code_or_reference, code_aliases)
        ),
        None,
    )
    qty_col = next(
        (
            column
            for header, column in headers.items()
            if _matches_header(header, config.columns.quantity, qty_aliases)
        ),
        None,
    )
    item_col = next(
        (
            column
            for header, column in headers.items()
            if config.columns.item and _matches_header(header, config.columns.item, item_aliases)
        ),
        None,
    )

    return [(item_col, code_col, qty_col)] if code_col and qty_col else []


def read_budget_items(excel_path: str | Path, config: ExcelConfig) -> list[BudgetItem]:
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {path}")

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active

    column_groups = _discover_column_groups(sheet, config)
    if not column_groups:
        headers_lidos = [
            _normalize_header(cell.value)
            for cell in sheet[config.header_row]
            if not _is_empty(cell.value)
        ]
        raise ValueError(
            "Não encontrei as colunas obrigatórias na planilha. "
            f"Colunas lidas normalizadas: {headers_lidos}. "
            "Verifique header_row e os nomes em config/config.yaml."
        )

    items: list[BudgetItem] = []
    for item_col, code_col, qty_col in column_groups:
        for row in range(config.first_data_row, sheet.max_row + 1):
            code_value = sheet.cell(row=row, column=code_col).value
            qty_value = sheet.cell(row=row, column=qty_col).value
            item_value = sheet.cell(row=row, column=item_col).value if item_col else None

            if _is_empty(code_value) and _is_empty(qty_value):
                continue
            if _is_empty(code_value):
                continue

            items.append(
                BudgetItem(
                    row_number=row,
                    code_or_reference=_to_clean_text(code_value),
                    quantity=qty_value if qty_value is not None else "",
                    item=_to_clean_text(item_value) if item_value is not None else None,
                )
            )

    workbook.close()
    return items


def unique_codes(items: Iterable[BudgetItem]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        code = item.code_or_reference.strip()
        if code and code not in seen:
            seen.add(code)
            result.append(code)
    return result


_CNPJ_VALUE_COLUMN = 17  # Coluna Q — célula mesclada Q:T onde fica o CNPJ do requisitante


def read_cnpj_from_excel(excel_path: str | Path, config: ExcelConfig) -> str | None:
    """Lê o CNPJ do requisitante da planilha.

    Varre as linhas antes do cabeçalho procurando a label 'CNPJ DO REQUISITANTE'
    e retorna o valor da coluna Q (17) da mesma linha, que é onde o CNPJ fica
    na célula mesclada Q:T.
    """
    path = Path(excel_path)
    if not path.exists():
        return None

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active

    for row_idx in range(1, config.header_row):
        for cell in sheet[row_idx]:
            label = _normalize_header(cell.value)
            if "CNPJ" in label:
                value_cell = sheet.cell(row=row_idx, column=_CNPJ_VALUE_COLUMN)
                value = _to_clean_text(value_cell.value)
                workbook.close()
                return value if value else None

    workbook.close()
    return None


def read_company_name_from_excel(excel_path: str | Path, config: ExcelConfig) -> str | None:
    """Lê o Nome Fantasia (ou Razão Social) do requisitante da planilha.

    Varre as linhas antes do cabeçalho procurando labels 'NOME FANTASIA' ou 'RAZAO SOCIAL'
    e retorna o valor da primeira célula não-vazia à direita do label encontrado.
    """
    path = Path(excel_path)
    if not path.exists():
        return None

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active

    for row_idx in range(1, config.header_row):
        for cell in sheet[row_idx]:
            label = _normalize_header(cell.value)
            if "FANTASIA" in label or "RAZAO SOCIAL" in label:
                # Procura a primeira célula não-vazia à direita do label
                for col_offset in range(1, 20):
                    value_cell = sheet.cell(row=row_idx, column=cell.column + col_offset)
                    value = _to_clean_text(value_cell.value)
                    if value:
                        workbook.close()
                        return value

    workbook.close()
    return None
