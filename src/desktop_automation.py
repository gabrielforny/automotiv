from __future__ import annotations

import re
import subprocess
import time
import webbrowser
from datetime import datetime
from pathlib import Path
from typing import Any

import pyautogui
import pyperclip
import yaml
from openpyxl import load_workbook
from pywinauto import Application, Desktop, keyboard

from src.config import BotConfig
from src.image_automation import ImageAutomation
from src.windows_tools import find_desktop_shortcut


class AutomotivApp:
    def __init__(self, config: BotConfig, logger: Any):
        self.config = config
        self.logger = logger
        self.app: Application | None = None
        self.image_config = self._load_image_config()
        self.image = ImageAutomation(
            assets_dir=self.image_config.get("assets_dir", config.images.assets_dir),
            logger=logger,
            confidence=float(self.image_config.get("confidence", config.images.confidence)),
        )
        pyautogui.PAUSE = config.runtime.pause_between_actions_seconds

    def start(self) -> None:
        shortcut = find_desktop_shortcut(self.config.app.shortcut_name)
        self.logger.info("Abrindo atalho: %s", shortcut)

        if self.config.runtime.dry_run:
            self.logger.info("DRY RUN: não vou abrir o sistema.")
            return

        subprocess.Popen([str(shortcut)], shell=True)
        self.app = Application(backend="uia").connect(
            title_re=self.config.app.window_title_regex,
            timeout=60,
        )
        self.logger.info("Sistema aberto/conectado.")

    def login(self, nome: str | None = None, senha: str | None = None) -> None:
        self.logger.info("Efetuando login.")
        if self.config.runtime.dry_run:
            return

        self._fill_login_fields(nome=nome, senha=senha)
        self.click_login()
        time.sleep(self.config.app.wait_after_login_seconds)

    def _fill_login_fields(self, nome: str | None = None, senha: str | None = None) -> None:
        time.sleep(5)
        login_nome = nome or self.config.app.login_password
        login_senha = senha or self.config.app.login_password

        if self._tentar_clicar_imagem("input_empresa_login", timeout=20):
            keyboard.send_keys("{TAB}")
            time.sleep(0.5)
            pyperclip.copy(login_nome)
            keyboard.send_keys("^a")
            keyboard.send_keys("^v")
            time.sleep(0.5)
            self.logger.info("Nome de login preenchido: %s", login_nome)
        else:
            self.logger.warning("Campo de nome não encontrado, pulando preenchimento do nome.")

        keyboard.send_keys("{TAB}")
        time.sleep(0.5)
        pyperclip.copy(login_senha)
        keyboard.send_keys("^a")
        keyboard.send_keys("^v")
        time.sleep(0.5)

    def click_login(self) -> None:
        self.logger.info("Clicando no botão LOGIN por imagem.")
        clicked = self._tentar_clicar_imagem("login_button", timeout=15)
        if not clicked:
            raise RuntimeError("Botão LOGIN não encontrado na tela. Verifique se o GRV está aberto e na tela de login.")
        self.logger.info("Clique no botão LOGIN realizado com sucesso por imagem.")

    def open_materials_screen(self) -> None:
        self.logger.info("Abrindo Cadastro > Materiais > Materiais/Itens.")
        if self.config.runtime.dry_run:
            return

        # Passo 1: clicar em Cadastro (imagem ou controle Windows)
        if not self._tentar_clicar_imagem("menu_cadastro", timeout=12):
            self.logger.info("menu_cadastro não encontrado por imagem. Tentando por controle Windows.")
            self._click_menu_by_control("Cadastro")
        time.sleep(0.8)

        # Passo 2: Materiais > Materiais/Itens (imagem ou teclado)
        if self._tentar_clicar_imagem("menu_materiais", timeout=12):
            time.sleep(0.8)
            if not self._tentar_clicar_imagem("menu_materiais_itens", timeout=12):
                self.logger.info("menu_materiais_itens não encontrado. Usando Down + Enter.")
                keyboard.send_keys("{DOWN}")
                time.sleep(0.2)
                keyboard.send_keys("{ENTER}")
        else:
            self.logger.info(
                "menu_materiais não encontrado por imagem. "
                "Navegando por teclado: Down×6, Right, Down, Enter."
            )
            for _ in range(6):
                keyboard.send_keys("{DOWN}")
                time.sleep(0.1)
            keyboard.send_keys("{RIGHT}")
            time.sleep(0.4)
            keyboard.send_keys("{DOWN}")
            time.sleep(0.2)
            keyboard.send_keys("{ENTER}")

        time.sleep(self.config.app.wait_after_open_screen_seconds)

    def open_clients_screen(self) -> None:
        self.logger.info("Abrindo Cadastro > Cliente.")
        if self.config.runtime.dry_run:
            return
        self._send_menu_sequence(["Cadastro"])
        keyboard.send_keys("{DOWN}")
        time.sleep(0.4)
        keyboard.send_keys("{DOWN}")
        time.sleep(0.4)
        keyboard.send_keys("{ENTER}")
        time.sleep(self.config.app.wait_after_open_screen_seconds)

    def abrir_novo_orcamento(self, company_name: str | None = None) -> None:
        """Menu: Orçamento > Orçamento > Novo (F2)."""
        self.logger.info("Abrindo Orçamento > Orçamento - F2/novo.")
        if self.config.runtime.dry_run:
            return
        self._send_menu_sequence(["Orçamento"])
        time.sleep(0.4)
        keyboard.send_keys("{DOWN}")
        time.sleep(0.4)
        keyboard.send_keys("{ENTER}")
        time.sleep(8)
        self._tentar_clicar_imagem("menu_orcamento_novo")
        time.sleep(2)
        if company_name:
            self.logger.info("Digitando nome do cliente no orçamento: %s", company_name)
            pyperclip.copy(company_name)
            keyboard.send_keys("^a")
            keyboard.send_keys("^v")
            time.sleep(0.3)
            keyboard.send_keys("{ENTER}")
            time.sleep(1)
        time.sleep(self.config.app.wait_after_open_screen_seconds)

    def open_previous_orders_search(self) -> None:
        """Abre Orçamento > Orçamento e clica em Pesquisar."""
        self.logger.info("Abrindo pesquisa de orçamentos anteriores.")
        if self.config.runtime.dry_run:
            return
        self._send_menu_sequence(["Orçamento"])
        time.sleep(0.4)
        keyboard.send_keys("{DOWN}")
        time.sleep(0.4)
        keyboard.send_keys("{ENTER}")
        time.sleep(8)
        self._tentar_clicar_imagem("search_f5_button", timeout=10)
        time.sleep(5)

    def _send_menu_sequence(self, labels: list[str]) -> None:
        submenu_open = False
        for label in labels:
            image_key = self._menu_label_to_image_key(label)
            image_name = self._get_image_name(image_key)
            if image_name:
                self.logger.info("Clicando menu '%s' por imagem '%s'.", label, image_name)
                self._tentar_clicar_imagem(image_key, timeout=12)
                time.sleep(0.8)
                submenu_open = True
                continue
            self.logger.info("Imagem do menu '%s' não configurada. Tentando por controle Windows.", label)
            self._click_menu_by_control(label, search_in_popup=submenu_open)
            time.sleep(0.8)
            submenu_open = True

    def _clicar_por_uia(
        self,
        title: str,
        control_type: str | None = None,
        timeout: int = 8,
        search_in_popup: bool = False,
    ) -> bool:
        """Clica num controle pelo título UIA. Retorna True se clicou, False se não achou."""
        desktop = Desktop(backend="uia")
        try:
            if search_in_popup:
                for win in desktop.windows():
                    try:
                        if not win.is_visible():
                            continue
                        kwargs = {"title": title}
                        if control_type:
                            kwargs["control_type"] = control_type
                        ctrl = win.child_window(**kwargs)
                        ctrl.wait("exists enabled visible", timeout=timeout)
                        ctrl.click_input()
                        return True
                    except Exception:
                        continue
                return False

            window = desktop.window(title_re=self.config.app.window_title_regex)
            window.set_focus()
            kwargs = {"title": title, "found_index": 0}
            if control_type:
                kwargs["control_type"] = control_type
            ctrl = window.child_window(**kwargs)
            ctrl.wait("exists enabled visible", timeout=timeout)
            ctrl.click_input()
            return True
        except Exception as exc:
            self.logger.debug("UIA click falhou para '%s': %s", title, exc)
            return False

    def _get_pesquisa_window(self, timeout: int = 8):
        """Retorna a janela 'Pesquisa' do GRV se estiver aberta, ou None."""
        desktop = Desktop(backend="uia")
        try:
            win = desktop.window(title="Pesquisa", class_name="TFMPesquisa_")
            win.wait("exists visible", timeout=timeout)
            return win
        except Exception:
            return None

    def _digitar_no_campo_pesquisa(self, text: str) -> bool:
        """Digita no campo 'Pesquisar:' da janela Pesquisa via UIA. Retorna True se conseguiu."""
        win = self._get_pesquisa_window()
        if win is None:
            return False
        try:
            campo = win.child_window(title="Pesquisar:", control_type="Pane").child_window(control_type="Edit")
            campo.wait("exists enabled visible", timeout=5)
            campo.set_focus()
            campo.set_edit_text(text)
            return True
        except Exception as exc:
            self.logger.debug("UIA: falhou ao digitar no campo Pesquisar: %s", exc)
            return False

    def _clicar_radio_pesquisa(self, title: str) -> bool:
        """Clica num RadioButton pelo título dentro da janela Pesquisa. Retorna True se clicou."""
        win = self._get_pesquisa_window()
        if win is None:
            return False
        try:
            radio = win.child_window(title=title, control_type="RadioButton")
            radio.wait("exists enabled visible", timeout=5)
            radio.click_input()
            return True
        except Exception as exc:
            self.logger.debug("UIA: falhou ao clicar radio '%s': %s", title, exc)
            return False

    def _click_menu_by_control(self, label: str, search_in_popup: bool = False) -> None:
        desktop = Desktop(backend="uia")
        window = desktop.window(title_re=self.config.app.window_title_regex)
        window.set_focus()
        try:
            if search_in_popup:
                # Após um clique de menu anterior, busca dentro do popup/submenu aberto
                try:
                    popup = desktop.window(control_type="Menu", visible_only=True)
                    control = popup.child_window(title=label, control_type="MenuItem")
                    control.wait("exists enabled visible", timeout=5)
                    control.click_input()
                    return
                except Exception:
                    pass  # fallback para busca na janela principal

            # Busca na janela toda, usando found_index=0 para evitar ambiguidade
            control = window.child_window(title=label, control_type="MenuItem", found_index=0)
            control.wait("exists enabled visible", timeout=8)
            control.click_input()
        except Exception as exc:
            self.logger.warning("Não consegui clicar no menu '%s' por controle: %s", label, exc)
            raise

    def press_search_f5(self) -> None:
        self.logger.info("Abrindo pesquisa com F5.")
        if self._clicar_por_uia("Pesquisa - F5", control_type="Button", timeout=5):
            self.logger.info("Botão 'Pesquisa - F5' clicado via UIA.")
        elif self._tentar_clicar_imagem("search_f5_button", timeout=5):
            self.logger.info("Botão de pesquisa clicado por imagem.")
        else:
            self.logger.info("Usando tecla F5.")
            keyboard.send_keys("{F5}")
        time.sleep(1)

    def close_current_screen(self) -> None:
        self.logger.info("Fechando tela atual.")
        if self.config.runtime.dry_run:
            return
        self._try_close_dialog()
        self._try_close_dialog()
        self._fechar_aba(timeout=10)
        time.sleep(1)

    def fechar_janela_para_buscar_site(self) -> None:
        self.logger.info("Fechando janela atual para liberar foco e buscar no site.")
        self._fechar_janela(timeout=10)
        time.sleep(3)
        self._fechar_janela(timeout=10)
        time.sleep(3)
        self._fechar_aba(timeout=10)
        time.sleep(3)

    def _image_esta_visivel(self, image_key: str, timeout: int = 3) -> bool:
        """Verifica se a imagem está na tela sem clicar. Retorna True se encontrou, False caso contrário."""
        image_name = self._get_image_name(image_key)
        if not image_name:
            return False
        return self.image.exists(image_name=image_name, timeout=timeout, confidence=self._get_confidence())

    def _tentar_clicar_imagem(self, image_key: str, timeout: int = 5, max_attempts: int = 3) -> bool:
        """Clica na imagem se ela aparecer. Tenta até max_attempts vezes antes de desistir."""
        image_name = self._get_image_name(image_key)
        if not image_name:
            self.logger.info("Imagem '%s' não configurada, pulando.", image_key)
            return False
        for attempt in range(1, max_attempts + 1):
            clicked = self.image.click(image_name=image_name, timeout=timeout, confidence=self._get_confidence())
            if clicked:
                if attempt > 1:
                    self.logger.info("Imagem '%s' encontrada na tentativa %s/%s.", image_key, attempt, max_attempts)
                return True
            if attempt < max_attempts:
                self.logger.info("Imagem '%s' não encontrada (tentativa %s/%s). Aguardando 1s...", image_key, attempt, max_attempts)
                time.sleep(1)
        self.logger.info("Imagem '%s' não encontrada após %s tentativas, pulando.", image_key, max_attempts)
        return False

    def _fechar_aba(self, timeout: int = 10) -> None:
        """Fecha a aba atual (MDI child). UIA: botão 'Sair-F10' na toolbar."""
        if self._clicar_por_uia("Sair-F10", control_type="Button", timeout=3):
            self.logger.info("Aba fechada via UIA ('Sair-F10').")
            return
        if not self._tentar_clicar_imagem("btn_fechar_aba", timeout=timeout):
            self.logger.info("btn_fechar_aba não encontrado por imagem. Usando ESC.")
            keyboard.send_keys("{ESC}")

    def _fechar_janela(self, timeout: int = 10) -> None:
        """Fecha popup/dialog aberto. UIA: botão '&Cancelar' ou 'Fechar' no popup."""
        desktop = Desktop(backend="uia")
        for win in desktop.windows():
            try:
                if not win.is_visible():
                    continue
                if re.search(r"CPS|GRV|AUTOMOTIV", win.window_text()):
                    continue
                rect = win.rectangle()
                if rect.width() < 50 or rect.height() < 50:
                    continue
                for title in ("&Cancelar", "Fechar"):
                    try:
                        btn = win.child_window(title=title, control_type="Button")
                        btn.wait("exists enabled visible", timeout=2)
                        btn.click_input()
                        self.logger.info("Popup fechado via UIA (botão '%s').", title)
                        return
                    except Exception:
                        continue
            except Exception:
                continue
        if not self._tentar_clicar_imagem("btn_fechar_janela", timeout=timeout):
            self.logger.info("btn_fechar_janela não encontrado por imagem. Usando F10.")
            keyboard.send_keys("{F10}")

    def _try_close_dialog(self) -> None:
        self._fechar_janela(timeout=10)
        time.sleep(0.8)
        keyboard.send_keys("{ESC}")
        time.sleep(0.8)

    def search_material(self, code_or_reference: str) -> bool:
        return self.find_material(code_or_reference=code_or_reference) is not None

    def find_material(self, code_or_reference: str) -> dict[str, Any] | None:
        """Pesquisa material e valida o código exato via Excel exportado da grid."""
        self.press_search_f5()
        self._select_search_by_code_internal()
        time.sleep(1)
        self._set_material_status()
        time.sleep(1)
        self._type_search_text(code_or_reference)
        time.sleep(3)

        if self._image_esta_visivel("sem_dados_na_busca", timeout=3):
            self.logger.info("Material %s não encontrado (sem dados na busca).", code_or_reference)
            return None

        exported_file = self._export_grid_to_excel(code_or_reference)
        try:
            material = self.find_material_in_exported_excel(exported_file, code_or_reference)
            if material:
                self.logger.info("Material encontrado no Excel exportado: %s", material)
                return material
            self.logger.info("Material não encontrado no Excel exportado: %s", code_or_reference)
            return None
        finally:
            self._delete_file_safely(exported_file)

    def _select_search_by_code_internal(self) -> None:
        # A lista "Pesquisar por" é um TcxGridSite (DevExpress) — não expõe linhas via UIA.
        # Estratégia: clicar no grid para focar, depois HOME para ir ao 1º item (Código Interno).
        # found_index=1 porque o índice 0 é a grid de resultados (maior, rect y>366).
        win = self._get_pesquisa_window()
        if win:
            try:
                grid = win.child_window(class_name="TcxGridSite", found_index=1)
                grid.wait("exists enabled visible", timeout=5)
                grid.click_input()
                keyboard.send_keys("{HOME}")
                self.logger.info("'Código Interno' selecionado via clique no grid + HOME.")
                return
            except Exception as exc:
                self.logger.debug("UIA: falhou ao focar grid Pesquisar por: %s", exc)
        if self._tentar_clicar_imagem("search_by_codigo_interno", timeout=10):
            return
        self.logger.info("search_by_codigo_interno não encontrado. Usando TAB+HOME.")
        keyboard.send_keys("{TAB}")
        keyboard.send_keys("{HOME}")

    def _set_material_status(self) -> None:
        if self._clicar_radio_pesquisa("Todos"):
            self.logger.info("Filtro 'Todos' selecionado via UIA.")
            time.sleep(0.3)
            return
        keyboard.send_keys("{TAB}{TAB}{DOWN}{DOWN}")
        time.sleep(0.3)

    def _type_search_text(self, text: str) -> None:
        self.logger.info("Digitando texto da pesquisa: %s", text)
        if self._digitar_no_campo_pesquisa(text):
            self.logger.info("Texto digitado via UIA no campo Pesquisar.")
            keyboard.send_keys("{ENTER}")
            return
        # Fallback teclado
        keyboard.send_keys("{TAB}")
        time.sleep(0.4)
        keyboard.send_keys("{TAB}")
        time.sleep(2)
        pyperclip.copy(text)
        time.sleep(2)
        keyboard.send_keys("^a")
        keyboard.send_keys("^v")
        time.sleep(2)
        keyboard.send_keys("{ENTER}")

    def _click_search_button(self) -> None:
        win = self._get_pesquisa_window()
        if win:
            try:
                btn = win.child_window(title="Pesquisar", control_type="Button")
                btn.wait("exists enabled visible", timeout=5)
                btn.click_input()
                self.logger.info("Botão 'Pesquisar' clicado via UIA.")
                return
            except Exception as exc:
                self.logger.debug("UIA: falhou ao clicar botão 'Pesquisar': %s", exc)
        if not self._tentar_clicar_imagem("search_button", timeout=5):
            self.logger.info("search_button não encontrado por imagem. Usando ENTER.")
            keyboard.send_keys("{ENTER}")

    def _export_grid_to_excel(self, expected_code: str, image_key: str = "export_excel_button") -> Path:
        self.logger.info("Exportando resultado da grid para Excel.")
        self._click_export_excel_button(image_key=image_key)
        export_path = self._save_exported_excel_dialog(expected_code)
        self._wait_until_file_is_ready(export_path, timeout=self.config.export.wait_timeout_seconds)
        self.logger.info("Arquivo exportado pronto em: %s", export_path)
        return export_path

    def _click_export_excel_button(self, image_key: str = "export_excel_button") -> None:
        self.logger.info("Clicando no botão de exportar Excel (imagem=%s).", image_key)
        # Modo âncora: localiza o ícone âncora e clica N pixels à direita para atingir o Excel
        if image_key == "export_excel_pedidos":
            anchor_name = self._get_image_name("export_excel_pedidos_anchor")
            if not anchor_name:
                raise RuntimeError("Imagem 'export_excel_pedidos_anchor' não configurada.")
            clicked = self.image.click_with_offset(
                image_name=anchor_name,
                offset_x=self.config.workflow.export_excel_pedidos_offset_x,
                timeout=10,
                confidence=self._get_confidence(),
            )
        else:
            image_name = self._get_image_name(image_key)
            if not image_name:
                raise RuntimeError(f"Imagem '{image_key}' não configurada.")
            clicked = self.image.click_bottommost(image_name=image_name, timeout=10, confidence=self._get_confidence())
        if not clicked:
            raise RuntimeError(f"Não consegui clicar no botão Excel ({image_key}).")
        time.sleep(1)

    def _save_exported_excel_dialog(self, expected_code: str, timeout: int = 15) -> Path:
        download_dir = Path(self.config.export.download_dir)
        download_dir.mkdir(parents=True, exist_ok=True)
        safe_code = self._sanitize_filename(expected_code)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = download_dir / f"{self.config.export.exported_file_prefix}_{safe_code}_{timestamp}.xlsx"
        self.logger.info("Salvando Excel exportado em: %s", file_path)

        # A janela Salvar como costuma vir com o campo Nome selecionado.
        # Colar caminho completo define pasta + nome em uma única ação.
        pyperclip.copy(str(file_path))
        keyboard.send_keys("^a")
        keyboard.send_keys("^v")
        time.sleep(0.5)
        keyboard.send_keys("{ENTER}")

        self._handle_save_confirmation_if_exists()
        return file_path

    def _handle_save_confirmation_if_exists(self) -> None:
        time.sleep(0.8)
        if not self._get_image_name("modal_duplicado") or not self._get_image_name("button_ok_duplicad"):
            return
        try:
            if self.image.exists(self._get_required_image_name("modal_duplicado"), timeout=2, confidence=self._get_confidence()):
                self._tentar_clicar_imagem("button_ok_duplicad", timeout=5)
        except Exception as exc:
            self.logger.info("Sem confirmação de substituição ou não consegui validar: %s", exc)

    def _wait_until_file_is_ready(self, file_path: Path, timeout: int = 90) -> None:
        self.logger.info("Aguardando arquivo ficar pronto: %s", file_path)
        deadline = time.time() + timeout
        last_size = -1
        stable_count = 0
        while time.time() < deadline:
            if file_path.exists():
                current_size = file_path.stat().st_size
                if current_size > 0 and current_size == last_size:
                    stable_count += 1
                else:
                    stable_count = 0
                last_size = current_size
                if stable_count >= 2:
                    try:
                        with file_path.open("rb"):
                            pass
                        return
                    except PermissionError:
                        self.logger.info("Arquivo ainda em uso. Aguardando...")
            time.sleep(1)
        raise TimeoutError(f"O arquivo exportado não ficou pronto dentro de {timeout} segundos: {file_path}")

    def find_material_in_exported_excel(self, excel_path: str | Path, expected_code: str) -> dict[str, Any] | None:
        path = Path(excel_path)
        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        headers = {str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value}
        self.logger.debug("Colunas no Excel de materiais: %s", list(headers.keys()))
        code_col = (
            headers.get("*Código Interno")
            or headers.get("Código Interno")
            or headers.get("*Codigo Interno")
            or headers.get("Codigo Interno")
        )
        item_col = headers.get("*Item") or headers.get("Item")
        family_col = headers.get("*Família") or headers.get("Família") or headers.get("*Familia") or headers.get("Familia")
        group_col = headers.get("*Grupo") or headers.get("Grupo")
        if not code_col:
            raise RuntimeError(
                f"Coluna 'Código Interno' não encontrada no Excel exportado. "
                f"Colunas disponíveis: {list(headers.keys())}"
            )

        expected = str(expected_code).strip()
        for row in range(2, sheet.max_row + 1):
            raw = sheet.cell(row=row, column=code_col).value
            if raw is None:
                continue
            # Valores numéricos do Excel chegam como float (ex: 47688952.0) — normaliza para string inteira
            if isinstance(raw, float) and raw == int(raw):
                code = str(int(raw))
            else:
                code = str(raw).strip()
            if code == expected:
                workbook.close()
                return {
                    "row": row,
                    "codigo_interno": code,
                    "item": sheet.cell(row=row, column=item_col).value if item_col else None,
                    "familia": sheet.cell(row=row, column=family_col).value if family_col else None,
                    "grupo": sheet.cell(row=row, column=group_col).value if group_col else None,
                }
        workbook.close()
        return None

    # ================================
    # Estrutura do fluxo de orçamento
    # ================================

    def criar_orcamento(
        self,
        company_code: str | None,
        items: list[Any],
        margins_by_code: dict[str, str] | None = None,
        carrier_code: str | None = None,
        company_name: str | None = None,
    ) -> str | None:
        self.logger.info("Iniciando criação de orçamento para empresa=%s | nome=%s | itens=%s", company_code, company_name, len(items))
        if self.config.runtime.dry_run:
            return None

        self._fechar_aba(timeout=10)
        time.sleep(3)

        self.abrir_novo_orcamento(company_name=company_name)

        time.sleep(5)
        self._clicar_campo_codigo_grade()
        time.sleep(2)

        total_items = len(items)
        for index, item in enumerate(items, start=1):
            code = getattr(item, "code_or_reference", None) or item.get("code_or_reference")
            quantity = getattr(item, "quantity", None) or item.get("quantity")
            margin = (margins_by_code or {}).get(str(code), self.config.workflow.default_margin)

            self.logger.info("Adicionando item no orçamento: código=%s | quantidade=%s | margem=%s", code, quantity, margin)
            self._digitar_codigo_item(str(code))
            self._digitar_quantidade_item(str(quantity))
            if margin:
                self._digitar_margem_item(str(margin))
            if index < total_items:
                self._ir_proxima_linha_grade(index=index)

        return self._preencher_dados_e_observacao(carrier_code=carrier_code)

    def _clicar_campo_codigo_grade(self) -> None:
        self._tentar_clicar_imagem("campo_codigo_grade_orcamento", timeout=10)
        time.sleep(3)
        try:
            self._tentar_clicar_imagem("lupa_dentro_selecao", timeout=10)
        except Exception as exc:
            self._tentar_clicar_imagem("lupa_fora_selecao", timeout=10)
        time.sleep(2)
        keyboard.send_keys("{ESC}")
        time.sleep(2)
        self.logger.warning("Clicado na coluna de código de grade")

    def _digitar_codigo_item(self, code: str) -> None:
        self.logger.warning("Digitando código do item: %s", code)
        pyautogui.write(code, interval=0.05)
        time.sleep(0.3)
        keyboard.send_keys("{ENTER}")
        time.sleep(0.5)
        self.handle_optional_message_modal()

    def _digitar_quantidade_item(self, quantity: str) -> None:
        self.logger.warning("Digitando quantidade do item: %s", quantity)
        keyboard.send_keys("{TAB}")
        time.sleep(0.4)
        keyboard.send_keys("{TAB}")
        time.sleep(0.5)
        keyboard.send_keys("{TAB}")
        time.sleep(0.5)
        keyboard.send_keys("{TAB}")
        time.sleep(0.5)
        keyboard.send_keys("{TAB}")
        time.sleep(0.5)
        keyboard.send_keys("{TAB}")
        time.sleep(0.5)
        keyboard.send_keys("{TAB}")
        pyperclip.copy(str(quantity))
        keyboard.send_keys("^a")
        keyboard.send_keys("^v")
        keyboard.send_keys("{ENTER}")
        time.sleep(0.3)

    def _digitar_margem_item(self, margin: str) -> None:
        keyboard.send_keys("{TAB}")
        time.sleep(0.4)
        keyboard.send_keys("{TAB}")
        time.sleep(0.5)
        keyboard.send_keys("{TAB}")
        time.sleep(0.5)
        keyboard.send_keys("{TAB}")
        time.sleep(0.5)
        keyboard.send_keys("{TAB}")
        time.sleep(0.5)
        keyboard.send_keys("{TAB}")
        time.sleep(0.5)
        keyboard.send_keys("{TAB}")
        time.sleep(0.5)
        keyboard.send_keys("{TAB}")
        time.sleep(0.5)
        keyboard.send_keys("{TAB}")
        pyperclip.copy(str(margin))
        keyboard.send_keys("^a")
        keyboard.send_keys("^v")
        keyboard.send_keys("{ENTER}")
        time.sleep(0.3)

    def _ir_proxima_linha_grade(self, index: int) -> None:
        self.logger.info("Indo para próxima linha do orçamento. Linha atual=%s", index)
        keyboard.send_keys("{DOWN}")
        time.sleep(0.4)

    def _preencher_dados_e_observacao(self, carrier_code: str | None = None) -> str | None:
        self.logger.info("Preenchendo dados finais do orçamento e observação. carrier_code=%r", carrier_code)

        self._tentar_clicar_imagem("aba_dados_orcamento", timeout=8)
        time.sleep(3)
        self._tentar_clicar_imagem("icone_obrigatorio", timeout=8)
        time.sleep(2)
        pyperclip.copy(self.config.workflow.observation_placeholder)
        keyboard.send_keys("^a")
        keyboard.send_keys("^v")
        time.sleep(1)
        keyboard.send_keys("{ENTER}")
        time.sleep(2)
        self.logger.info("Preenchido campo de dados do orçamento.")
        self._tentar_clicar_imagem("aba_observacao_orcamento", timeout=8)
        time.sleep(2)
        self._tentar_clicar_imagem("icone_obrigatorio", timeout=8)
        time.sleep(2)
        observation = self._montar_texto_observacao(carrier_code=carrier_code)
        self.logger.info("Texto da observação a ser preenchida:\n%s", observation)
        time.sleep(1)
        pyperclip.copy(observation)
        keyboard.send_keys("^a")
        keyboard.send_keys("^v")
        self.logger.info("Preenchido campo de observação do orçamento.")
        time.sleep(1)
        keyboard.send_keys("{ENTER}")
        return None

    def _montar_texto_observacao(self, carrier_code: str | None) -> str:
        if not carrier_code:
            return self.config.workflow.observation_placeholder
        return carrier_code

    def buscar_margens_pedidos_anteriores(
        self,
        company_code: str | None,
        material_items: dict[str, str | None],  # {code: descricao_ou_none}
        company_name: str | None = None,
    ) -> tuple[dict[str, str], str | None]:
        """Busca margens e transportadora em orçamentos anteriores do cliente.

        Retorna (margins_by_code, carrier_code).
        """
        self.logger.info(
            "Buscando margens anteriores para empresa=%s | nome=%s | itens=%s",
            company_code, company_name, list(material_items.keys()),
        )
        if not company_code or self.config.runtime.dry_run:
            return {}, None

        self.open_previous_orders_search()
        self._filtrar_pedidos_anteriores_por_cliente()

        time.sleep(3)
        if self._image_esta_visivel("sem_dados_na_busca", timeout=3):
            self.logger.info("Nenhum pedido anterior para o cliente %s (sem dados na busca).", company_code)
            self._fechar_tela_pedidos_anteriores()
            return {}, None

        exported_list = self._export_grid_to_excel(company_code, image_key="export_excel_pedidos")
        try:
            pedidos = self._ler_pedidos_do_excel_exportado(exported_list, company_code)
        finally:
            self._delete_file_safely(exported_list)

        if not pedidos:
            self.logger.info("Nenhum pedido anterior encontrado para o cliente %s.", company_code)
            self._fechar_tela_pedidos_anteriores()
            return {}, None
        
        margins: dict[str, str] = {}
        carrier_code: str | None = None

        # Fecha a janela de resultados da exportação para voltar ao form de pesquisa
        self._fechar_janela(timeout=10)
        time.sleep(1)

        for order_idx, (n_orcamento, _) in enumerate(pedidos):
            target_remaining = {c for c in material_items if c not in margins}
            if not target_remaining and carrier_code:
                break

            self.logger.info(
                "Pedido %s/%s: N=%s | itens restantes=%s",
                order_idx + 1, len(pedidos),
                n_orcamento, target_remaining,
            )

            found_carrier = self._buscar_margens_em_pedido(
                n_orcamento=n_orcamento,
                material_items=material_items,
                margins=margins,
            )

            if found_carrier and not carrier_code:
                carrier_code = found_carrier

        self._fechar_tela_pedidos_anteriores()
        return margins, carrier_code

    def _ler_pedidos_do_excel_exportado(
        self, excel_path: Path, company_code: str
    ) -> list[tuple[str, str | None]]:
        """Lê o Excel exportado de orçamentos, filtra pelo Cód. Cliente e ordena por data desc.

        Colunas por posição (1-indexed):
          1  = N orçamento
          11 = Código da Transportadora
          24 = Data Orçamento  (anterior ao Cód. Cliente)
          25 = Cód. Cliente (filtro)
        """
        workbook = load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        cod_alvo = str(company_code).strip()

        rows_raw: list[tuple[str, str | None, any]] = []
        for row in range(2, sheet.max_row + 1):
            cod_cliente = str(sheet.cell(row=row, column=25).value or "").strip()
            if cod_cliente != cod_alvo:
                continue
            n_orcamento = str(sheet.cell(row=row, column=1).value or "").strip()
            if not n_orcamento:
                continue
            transportadora = str(sheet.cell(row=row, column=11).value or "").strip() or None
            data_orcamento = sheet.cell(row=row, column=24).value  # datetime ou string
            rows_raw.append((n_orcamento, transportadora, data_orcamento))

        workbook.close()

        # Ordenar por data decrescente (mais recente primeiro); None vai para o fim
        def _sort_key(item):
            data = item[2]
            if data is None:
                return (0, "")
            if hasattr(data, "toordinal"):  # datetime
                return (1, data.toordinal())
            return (1, str(data))

        rows_raw.sort(key=_sort_key, reverse=True)
        pedidos = [(n, t) for n, t, _ in rows_raw]

        self.logger.info(
            "Pedidos do cliente %s: %s encontrado(s) (ordenados por data desc): %s",
            company_code, len(pedidos), [p[0] for p in pedidos],
        )
        return pedidos

    def _buscar_margens_em_pedido(
        self,
        n_orcamento: str,
        material_items: dict[str, str | None],
        margins: dict[str, str],
    ) -> str | None:
        """Abre um pedido anterior, pesquisa cada item pelo campo de busca e extrai margens.

        Muta `margins` com os valores encontrados.
        Retorna o código da transportadora lido do orçamento, ou None.
        """
        self.logger.info("Processando pedido N=%s.", n_orcamento)
        target_codes = {c for c in material_items if c not in margins}

        # Passo 1: garantir que estamos no form de pesquisa (filtrar_por_codigo visível)
        if not self._image_esta_visivel("filtrar_por_codigo", timeout=3):
            self.logger.info("filtrar_por_codigo não visível — clicando search_f5_button para voltar ao form.")
            self._tentar_clicar_imagem("limpar_pesquisa_item_orcamento", timeout=5)
            time.sleep(0.5)
            self._tentar_clicar_imagem("search_f5_button", timeout=10)
            time.sleep(1)

        self._tentar_clicar_imagem("filtrar_por_codigo", timeout=10)
        time.sleep(1)

        # Passo 2: TAB × 3 para chegar ao campo do N do orçamento
        for _ in range(3):
            keyboard.send_keys("{TAB}")
            time.sleep(0.2)

        # Passo 3: digitar N do orçamento
        pyperclip.copy(str(n_orcamento))
        keyboard.send_keys("^a")
        keyboard.send_keys("^v")
        time.sleep(0.5)

        # Passo 4: pesquisar e verificar resultado
        keyboard.send_keys("{ENTER}")
        time.sleep(3)

        if self._image_esta_visivel("sem_dados_na_busca", timeout=3):
            self.logger.info("Pedido %s não encontrado na busca. Pulando.", n_orcamento)
            self._tentar_clicar_imagem("search_f5_button", timeout=10)
            time.sleep(1)
            return None

        # Passo 5: abrir o único resultado com ENTER
        keyboard.send_keys("{ENTER}")
        time.sleep(3)

        # Passo 7: pesquisar cada item pelo campo de busca dentro do orçamento
        carrier_found: str | None = None
        _sentinel = "__LENDO__"
        codes_to_search = list(target_codes)

        # Primeira pesquisa: clicar no campo de busca de item
        self._tentar_clicar_imagem("input_pesquisa_item_orcamento", timeout=10)
        time.sleep(0.5)

        for code_idx, code in enumerate(codes_to_search):
            if code in margins:
                # Item já encontrado em iteração anterior; limpar se houver próximo
                has_next = any(c not in margins for c in codes_to_search[code_idx + 1:])
                if has_next:
                    self._tentar_clicar_imagem("limpar_pesquisa_item_orcamento", timeout=5)
                    time.sleep(0.5)
                continue

            # Campo já está focado (inicial ou pós-limpar)
            pyperclip.copy(code)
            keyboard.send_keys("^a")
            keyboard.send_keys("^v")
            time.sleep(0.5)
            keyboard.send_keys("{ENTER}")
            time.sleep(1.5)

            self.logger.info("Pedido %s | buscando código=%s", n_orcamento, code)

            lupa_dentro = self._image_esta_visivel("lupa_dentro_selecao", timeout=3)
            lupa_fora = not lupa_dentro and self._image_esta_visivel("lupa_fora_selecao", timeout=3)

            if lupa_dentro or lupa_fora:
                # Clicar na lupa que apareceu
                if not self._tentar_clicar_imagem("lupa_dentro_selecao", timeout=3):
                    self._tentar_clicar_imagem("lupa_fora_selecao", timeout=3)
                time.sleep(0.5)

                # TAB × 17 → ENTER → Ctrl+C para ler Margem de Lucro
                for _ in range(17):
                    keyboard.send_keys("{TAB}")
                    time.sleep(0.1)
                keyboard.send_keys("{ENTER}")
                time.sleep(0.3)
                pyperclip.copy(_sentinel)
                time.sleep(0.1)
                keyboard.send_keys("^c")
                time.sleep(0.3)
                margin = pyperclip.paste().strip()

                if margin and margin != _sentinel:
                    margins[code] = margin
                    self.logger.info("Margem extraída: código=%s | margem=%s", code, margin)
                else:
                    self.logger.warning("Pedido %s | código=%s: margem ilegível.", n_orcamento, code)

                # Ler transportadora (apenas na primeira vez que encontramos um item)
                if not carrier_found:
                    self._tentar_clicar_imagem("aba_transporte_orcamento", timeout=5)
                    time.sleep(0.5)
                    for _ in range(2):
                        keyboard.send_keys("{TAB}")
                        time.sleep(0.2)
                    pyperclip.copy(_sentinel)
                    time.sleep(0.1)
                    keyboard.send_keys("^c")
                    time.sleep(0.3)
                    carrier_raw = pyperclip.paste().strip()
                    if carrier_raw and carrier_raw != _sentinel:
                        carrier_found = carrier_raw
                        self.logger.info("Transportadora lida: %s", carrier_found)

                    # Voltar para aba Itens
                    self._tentar_clicar_imagem("aba_itens_orcamento", timeout=5)
                    time.sleep(0.5)
            else:
                self.logger.info("Pedido %s | código=%s: não encontrado neste orçamento.", n_orcamento, code)

            # Se há mais itens para buscar, limpar o campo de pesquisa
            has_next = any(c not in margins for c in codes_to_search[code_idx + 1:])
            if has_next:
                self._tentar_clicar_imagem("limpar_pesquisa_item_orcamento", timeout=5)
                time.sleep(0.5)
                # Após limpar, o campo já está focado — não precisa clicar em input novamente

        # Voltar ao form de pesquisa para a próxima iteração
        self._tentar_clicar_imagem("search_f5_button", timeout=10)
        time.sleep(1)

        return carrier_found

    def _filtrar_pedidos_anteriores_por_cliente(self) -> None:
        self._tentar_clicar_imagem("btn_limpar_filtro_ja_feito", timeout=8)
        time.sleep(2)
        keyboard.send_keys("{ENTER}")
        time.sleep(2)
        self._tentar_clicar_imagem("confirmar_pesquisa_muitos_itens", timeout=8)
        time.sleep(0.5)
        self._tentar_clicar_imagem("click_ok_pesquisa_itens", timeout=8)
        time.sleep(7)

    def _calcular_data_inicio_pesquisa(self) -> str:
        from datetime import date
        months_back = self.config.workflow.previous_orders_months_back
        today = date.today()
        month = today.month - (months_back % 12)
        year = today.year - (months_back // 12)
        if month <= 0:
            month += 12
            year -= 1
        return f"{today.day:02d}/{month:02d}/{year}"

    def _contar_linhas_excel(self, excel_path: Path) -> int:
        workbook = load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        count = max(0, sheet.max_row - 1)  # -1 para ignorar o cabeçalho
        workbook.close()
        return count

    def _fechar_tela_pedidos_anteriores(self) -> None:
        self._try_close_dialog()
        self._fechar_aba(timeout=10)
        time.sleep(0.8)

    def buscar_codigo_transportadora(self, company_code: str | None) -> str | None:
        """Busca o código da transportadora do último pedido ou aba PADRÕES do cadastro do cliente."""
        self.logger.info("Buscando código da transportadora para empresa=%s", company_code)
        if not company_code or self.config.runtime.dry_run:
            return None
        # TODO: Implementar após capturar telas de pedidos anteriores ou aba PADRÕES do cliente.
        return None

    def gravar_orcamento(self) -> None:
        self.logger.info("Gravando orçamento.")
        self._tentar_clicar_imagem("botao_gravar_f3", timeout=10)
        time.sleep(1)
        self._tratar_erro_gravacao(True)
        time.sleep(2)

    def _tratar_erro_gravacao(self, is_gravar: bool) -> None:
        for attempt in range(1, self.config.workflow.save_retry_attempts + 1):
            self.logger.info("Verificando erro de gravação. Tentativa %s", attempt)
            self._tentar_clicar_imagem("save_error_modal_ok_button", timeout=10)
            time.sleep(3)
            self._tentar_clicar_imagem("save_error_modal_ok_button", timeout=10)
            time.sleep(0.5)
            if is_gravar:
                self.logger.info("Tentando gravar novamente após erro.")
                self.gravar_orcamento()

    def gerar_pdf_orcamento(self) -> None:
        self.logger.info("Gerando PDF do orçamento.")
        self._send_menu_sequence(["Imprimir", "Imprimir Orçamento Padrão"])
        time.sleep(2)
        self._tratar_erro_gravacao(False)
        time.sleep(2)
        self._tentar_clicar_imagem("print_ok_button", timeout=10)

    # ================================
    # Cliente
    # ================================

    def open_fallback_site(self) -> None:
        self.logger.info("Abrindo site fallback: %s", self.config.search.fallback_site)
        if not self.config.runtime.dry_run:
            webbrowser.open(self.config.search.fallback_site)

    def search_client_and_get_code(self, cpf_or_cnpj: str) -> tuple[str | None, str | None, str | None]:
        """Retorna (client_code, carrier_code, company_name). Carrier lido da aba PADRÕES."""
        self.logger.info("Pesquisando cliente por CPF/CNPJ: %s", cpf_or_cnpj)
        if self.config.runtime.dry_run:
            return None, None, None
        self.press_search_f5()
        self._select_search_by_cpf_cnpj()
        self._garantir_radio_todos()
        self._type_search_text(cpf_or_cnpj)
        time.sleep(2)
        code, company_name = self._extract_first_client_code(cpf_or_cnpj)
        carrier_code = None
        if code:
            self._fechar_confirmacao_exportacao()
            carrier_code = self._ler_codigo_transportadora_aba_padrao()
        self._fechar_aba(timeout=5)
        return code, carrier_code, company_name

    def _fechar_confirmacao_exportacao(self) -> None:
        """Fecha diálogo de exportação e confirma seleção do cliente via &OK na janela Pesquisa."""
        self._fechar_janela(timeout=5)
        time.sleep(5)
        win = self._get_pesquisa_window()
        if win:
            try:
                btn = win.child_window(title="&OK", control_type="Button")
                btn.wait("exists enabled visible", timeout=5)
                btn.click_input()
                self.logger.info("Seleção confirmada via UIA ('&OK' na Pesquisa).")
                time.sleep(0.5)
                return
            except Exception as exc:
                self.logger.debug("UIA: falhou &OK na Pesquisa: %s", exc)
        if not self._tentar_clicar_imagem("btn_ok_exportacao", timeout=5):
            keyboard.send_keys("{ENTER}")
        time.sleep(0.5)

    def _ler_codigo_transportadora_aba_padrao(self) -> str | None:
        """Clica na aba PADRÕES do cadastro do cliente, navega até o campo de transportadora e copia."""
        self._tentar_clicar_imagem("aba_padrao_cliente", timeout=8)
        time.sleep(0.5)
        self._tentar_clicar_imagem("interrogacao_aba_padrao", timeout=8)
        time.sleep(0.5)
        keyboard.send_keys("{ESC}")
        time.sleep(0.5)
        for _ in range(17):
            keyboard.send_keys("{TAB}")
            time.sleep(0.1)
        time.sleep(0.5)
        keyboard.send_keys("^a")
        keyboard.send_keys("^c")
        time.sleep(0.3)
        value = pyperclip.paste().strip()
        self.logger.info("Código de transportadora da aba PADRÕES: %s", value or "(vazio)")
        return value or None

    def _select_search_by_cpf_cnpj(self) -> None:
        self._tentar_clicar_imagem("search_by_cnpj_cpf", timeout=10)
        time.sleep(0.5)

    def _garantir_radio_contendo(self) -> None:
        if self._clicar_radio_pesquisa("Contendo"):
            self.logger.info("Filtro 'Contendo' selecionado via UIA.")
            time.sleep(0.3)
            return
        if self._tentar_clicar_imagem("radio_contendo", timeout=5):
            self.logger.info("Filtro 'Contendo' selecionado por imagem.")
            time.sleep(0.3)
        else:
            self.logger.warning("Filtro 'Contendo' não encontrado. Seguindo sem alterar o filtro.")

    def _garantir_radio_todos(self) -> None:
        if self._clicar_radio_pesquisa("Todos"):
            self.logger.info("Filtro 'Todos' selecionado via UIA.")
            time.sleep(0.3)
            return
        if self._tentar_clicar_imagem("radio_todos", timeout=5):
            self.logger.info("Filtro 'Todos' selecionado por imagem.")
            time.sleep(0.3)
        else:
            self.logger.warning("Filtro 'Todos' não encontrado. Seguindo sem alterar o filtro.")

    def _extract_first_client_code(self, cpf_or_cnpj: str) -> tuple[str | None, str | None]:
        """Retorna (client_code, company_name) da primeira linha da grid de clientes exportada."""
        self.logger.info("Exportando grid de clientes para Excel.")
        exported_file = self._export_grid_to_excel(cpf_or_cnpj)
        try:
            code, name = self._find_client_info_in_exported_excel(exported_file)
            if code:
                self.logger.info("Cliente extraído: código=%s | nome=%s", code, name)
            else:
                self.logger.info("Nenhum código de cliente encontrado no Excel exportado.")
            return code, name
        finally:
            self._delete_file_safely(exported_file)

    def _find_client_info_in_exported_excel(self, excel_path: Path) -> tuple[str | None, str | None]:
        """Retorna (client_code, company_name) da primeira linha do Excel exportado de clientes."""
        workbook = load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        headers = {str(cell.value or "").strip(): cell.column for cell in sheet[1] if cell.value}
        self.logger.info("Colunas no Excel de clientes exportado: %s", list(headers.keys()))
        code_col = headers.get("*Código") or headers.get("Código") or headers.get("*Codigo") or headers.get("Codigo")
        if code_col is None:
            raise RuntimeError(
                f"Coluna '*Código' não encontrada no Excel de clientes exportado. "
                f"Colunas disponíveis: {list(headers.keys())}"
            )
        name_col = (
            headers.get("*Nome Fantasia")
            or headers.get("Nome Fantasia")
            or headers.get("*Razão Social")
            or headers.get("Razão Social")
            or headers.get("*Razao Social")
            or headers.get("Razao Social")
            or headers.get("*Nome")
            or headers.get("Nome")
        )
        for row in range(2, sheet.max_row + 1):
            code_value = sheet.cell(row=row, column=code_col).value
            if code_value is not None and str(code_value).strip():
                code = str(code_value).strip()
                name: str | None = None
                if name_col:
                    nv = sheet.cell(row=row, column=name_col).value
                    name = str(nv).strip() if nv is not None and str(nv).strip() else None
                workbook.close()
                return code, name
        workbook.close()
        return None, None

    # ================================
    # Controle do ciclo de vida da app
    # ================================

    def close_app(self) -> None:
        """Fecha a janela do GRV para liberar o foco antes de abrir o navegador."""
        self.logger.info("Fechando aplicação GRV.")
        self._fechar_janela(timeout=10)
        time.sleep(1)
        self._fechar_janela(timeout=10)
        time.sleep(1)
        self._fechar_aba(timeout=10)
        time.sleep(3)
        self._tentar_clicar_imagem("btn_sair_grv", timeout=10)
        time.sleep(1)
        self._tentar_clicar_imagem("btn_sair_do_sistema", timeout=10)
        time.sleep(3)
        self.app = None

    def reopen(self) -> None:
        """Reabre e faz login no GRV após ter fechado para ir ao site."""
        self.logger.info("Reabrindo aplicação GRV.")
        self.start()
        self.login()

    def focus_app(self) -> None:
        """Traz a janela do GRV de volta ao foco após o browser fechar."""
        self.logger.info("Trazendo GRV de volta ao foco.")
        if self.config.runtime.dry_run:
            return
        try:
            desktop = Desktop(backend="uia")
            window = desktop.window(title_re=self.config.app.window_title_regex)
            window.set_focus()
            time.sleep(0.5)
        except Exception as exc:
            self.logger.warning("Não foi possível focar a janela do GRV: %s", exc)

    # ================================
    # Utilitários
    # ================================

    def handle_optional_message_modal(self) -> None:
        self.logger.info("Verificando se apareceu modal de mensagem.")
        time.sleep(0.5)
        modal_image = self._get_image_name("message_modal")
        ok_image = self._get_image_name("message_modal_ok_button")
        if not modal_image or not ok_image:
            self.logger.info("Imagens do modal não configuradas. Seguindo fluxo normal.")
            return
        modal_appeared = self.image.exists(image_name=modal_image, timeout=3, confidence=self._get_confidence())
        if not modal_appeared:
            self.logger.info("Modal de mensagem não apareceu. Seguindo fluxo normal.")
            return
        self.logger.info("Modal de mensagem encontrado. Clicando no OK.")
        clicked_ok = self.image.click(image_name=ok_image, timeout=5, confidence=self._get_confidence())
        if not clicked_ok:
            raise RuntimeError("Modal apareceu, mas não consegui clicar no botão OK.")
        time.sleep(0.5)

    def _click_configured_image_or_fail(self, image_key: str, timeout: int = 10, confidence: float | None = None) -> None:
        image_name = self._get_image_name(image_key)
        if not image_name:
            raise RuntimeError(f"Imagem '{image_key}' não configurada em config/config.yaml na seção images.")
        clicked = self.image.click(image_name=image_name, timeout=timeout, confidence=confidence or self._get_confidence())
        if clicked:
            return
        assets_dir = self.image_config.get("assets_dir", self.config.images.assets_dir)
        raise RuntimeError(
            f"Não foi possível localizar/clicar na imagem '{image_key}' ({image_name}). "
            f"Verifique se o arquivo existe em '{assets_dir}', se o recorte está correto "
            "e se a tela está na mesma escala/resolução."
        )

    def _get_image_name(self, image_key: str) -> str | None:
        value = self.image_config.get(image_key)
        if value is None or str(value).strip() == "":
            return None
        return str(value).strip()

    def _get_required_image_name(self, image_key: str) -> str:
        image_name = self._get_image_name(image_key)
        if not image_name:
            raise RuntimeError(f"Imagem obrigatória '{image_key}' não configurada em config/config.yaml.")
        return image_name

    def _get_confidence(self) -> float:
        return float(self.image_config.get("confidence", self.config.images.confidence))

    def _load_image_config(self) -> dict[str, Any]:
        return self.config.images.model_dump()

    @staticmethod
    def _menu_label_to_image_key(label: str) -> str:
        normalized = (
            label.lower()
            .replace("/", "_")
            .replace(" ", "_")
            .replace("-", "_")
            .replace("á", "a")
            .replace("à", "a")
            .replace("â", "a")
            .replace("ã", "a")
            .replace("é", "e")
            .replace("ê", "e")
            .replace("í", "i")
            .replace("ó", "o")
            .replace("ô", "o")
            .replace("õ", "o")
            .replace("ú", "u")
            .replace("ç", "c")
        )
        mapping = {
            "cadastro": "menu_cadastro",
            "cliente": "menu_cliente",
            "materiais": "menu_materiais",
            "materiais_itens": "menu_materiais_itens",
            "orcamento": "menu_orcamento",
            "orcamento_submenu": "menu_orcamento_submenu",
            "orcamento_novo": "menu_orcamento_novo",
            "orcamento_pesquisar": "menu_orcamento_pesquisar",
            "imprimir": "menu_imprimir",
            "imprimir_orcamento_padrao": "menu_imprimir_orcamento_padrao",
        }
        return mapping.get(normalized, f"menu_{normalized}")

    @staticmethod
    def _sanitize_filename(value: str) -> str:
        return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value).strip()) or "sem_codigo"

    def _delete_file_safely(self, path: Path) -> None:
        try:
            path.unlink(missing_ok=True)
            self.logger.info("Arquivo temporário removido: %s", path)
        except Exception as exc:
            self.logger.warning("Não consegui remover arquivo temporário %s: %s", path, exc)
